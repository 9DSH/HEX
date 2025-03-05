import gspread
import os
from datetime import datetime
from oauth2client.service_account import ServiceAccountCredentials  #pip install gspread oauth2client pandas
import pandas as pd
import logging
from googleapiclient.discovery import build  # pip install google-api-python-client google-auth google-auth-oauthlib google-auth-httplib2
from gspread_formatting import format_cell_range, CellFormat, Color  # Ensure gspread-formatting is installed
import tempfile
from openpyxl import Workbook
import time


# Logging configuration
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class GoogleManager:
    def __init__(self, creds_path):
        # Initialize the Google Sheets API client
        self.scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
        self.creds = ServiceAccountCredentials.from_json_keyfile_name(creds_path, self.scope)
        self.client = gspread.authorize(self.creds)
        self.last_shared_month = self.load_last_shared_month()  # Load the last shared month from a file
        self.is_uploading = True


    def create_monthly_spreadsheets(self):
        """Create a spreadsheet for each month of the year if it doesn't already exist."""
        months = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
        year = datetime.now().year  # Use the current year

        for month in months:
            spreadsheet_name = f"Report_{month}_{year}"
            
            # Check if the spreadsheet already exists
            spreadsheet_id = self.get_spreadsheet_id(spreadsheet_name)
            if spreadsheet_id:
                logger.info(f"Spreadsheet for {spreadsheet_name} already exists. Skipping creation.")
            else:
                # If the spreadsheet doesn't exist, create it
                self.create_spreadsheet(spreadsheet_name)

    def create_spreadsheet(self, spreadsheet_name):
        """Create a new spreadsheet with the given name."""
        try:
            service = build('sheets', 'v4', credentials=self.creds)
            spreadsheet = {
                'properties': {'title': spreadsheet_name}
            }
            request = service.spreadsheets().create(body=spreadsheet)
            response = request.execute()
            spreadsheet_id = response['spreadsheetId']
            logger.info(f"Created new spreadsheet: {spreadsheet_name}")
            self.share_spreadsheet(spreadsheet_id, "Hamed.dsh.ac@gmail.com")
        except Exception as e:
            logger.error(f"Error creating spreadsheet '{spreadsheet_name}': {str(e)}")

    def load_last_shared_month(self):
        """Load only the last shared month from a text file."""
        if os.path.exists('last_shared_month.txt'):
            with open('last_shared_month.txt', 'r') as file:
                lines = file.read().splitlines()  # Ensures we get the last non-empty line
                return lines[-1] if lines else ""  # Get last valid line
        return ""

    def save_last_shared_month(self, current_month):
        """Save the last shared month by appending to a text file."""
        with open('last_shared_month.txt', 'a') as file:  # Open in append mode
            file.write(current_month + '\n')  # Add a new line after writing
       
    def share_spreadsheet(self, spreadsheet_id, email_addresses):
        """Share the spreadsheet with the given list of email addresses."""
        service = build('drive', 'v3', credentials=self.creds)

        for email in email_addresses:  # Iterate over each email
            permission = {
                'type': 'user',
                'role': 'writer',  # Change as needed: 'reader', 'commenter', 'writer'
                'emailAddress': email
            }

            try:
                # Make the request to share
                service.permissions().create(
                    fileId=spreadsheet_id,
                    body=permission,
                    fields='id'
                ).execute()
                logger.info(f"Shared '{spreadsheet_id}' with {email}.")
            except Exception as e:
                logger.error(f"Failed to share '{spreadsheet_id}' with {email}: {e}")

    def get_spreadsheet_id(self, spreadsheet_name):
        """Get a spreadsheet ID by searching for its name."""
        try:
            # List all spreadsheets available in the account
            spreadsheet_list = self.client.list_spreadsheet_files()
            for spreadsheet in spreadsheet_list:
                if spreadsheet['name'] == spreadsheet_name:
                    return spreadsheet['id']
        except Exception as e:
            logger.error(f"Error searching for spreadsheet: {str(e)}")
        return None
    
    
    def ensure_monthly_spreadsheet_exists(self, current_month_year):
        """Ensure that a spreadsheet for the current month exists, returning its ID."""
        # Get the name for the current month
        spreadsheet_id = self.get_spreadsheet_id(current_month_year)  # Check for existing spreadsheet
        if not spreadsheet_id:  # If it doesn't exist, create a new one
            logger.info(f"No existing spreadsheet found for: {current_month_year}. Creating new one.")  

        return spreadsheet_id  # Return the ID of the spreadsheet
 
    
    def upload_csv(self, csv_file_path ,sheet_name, new_column_names):
        try:

            self.is_uploading = True
            # Open the Google Sheet
            # Get current month and year for naming
            current_month_year = datetime.now().strftime("Report_%B_%Y")
            # Ensure the monthly spreadsheet exists and fetch its ID
            spreadsheet_id = self.ensure_monthly_spreadsheet_exists(current_month_year)

            spreadsheet = self.client.open_by_key(spreadsheet_id)
            try:
                worksheet = spreadsheet.worksheet(sheet_name)
            except gspread.exceptions.WorksheetNotFound:
                worksheet = spreadsheet.add_worksheet(title=sheet_name, rows="500", cols="20")  # create a new sheet
                try:
                   default_sheet = spreadsheet.worksheet("Sheet1")
                   spreadsheet.del_worksheet(default_sheet)
                   logger.info("Deleted the default 'Sheet1'")
                except gspread.exceptions.WorksheetNotFound:
                   logger.info("No default 'Sheet1' found, nothing to delete.")
    
                      
            data = pd.read_csv(csv_file_path)

            # Sort the Hex Dashboard dataset by DATE
            # Check which columns apply based on the sheet name
            if sheet_name != "Clients": 
                logger.info("Sorting dataset by DATE")
                if sheet_name == "Transaction History":
                    date_column_name = 'transaction_date' 
                elif sheet_name == "Hex Dashboard":
                    date_column_name = 'DATE'
                elif sheet_name == "OrdersHistory":
                    date_column_name = 'Order_date'
                # Sort DataFrame by date
                sorted_data = data.sort_values(by=date_column_name, ascending=False).reset_index(drop=True)
                data = sorted_data


            # Clear the existing content
             # Define ranges based on sheet names
            ranges = {
                    "Hex Dashboard": 'A1:D33',
                    "Clients": 'A1:E500',
                    "OrdersHistory": 'A1:L500',
                    "Transaction History": 'A1:G500'
            }

            # Clear the specific range based on the sheet name
            if sheet_name in ranges:
                worksheet.batch_clear([ranges[sheet_name]])


            # Update the Google Sheet with the CSV data
            worksheet.update([data.columns.values.tolist()] + data.values.tolist())
           # If new column names are provided, update the column names in the sheet
            if new_column_names:
                worksheet.update('A1:' + chr(65 + len(new_column_names) - 1) + '1', [new_column_names])  # Update the header row

            # Define the columns to format
            numeric_format_columns = {
                'transaction': ['transaction_size'],  # Specific to transaction CSV
                'hex': ['Bought_USDT', 'Sold_USDT', 'Net_position'],  # Specific to hex CSV
                'client': ['USDT_Balance', 'Toman_Balance'],  # Specific to client CSV
                'order': ['Order_size', 'Order_price', 'Payable_to_Toman', 'paid_by_client']  # Specific to order CSV
            }
            
            

            # Columns to format based on the sheet name
            if sheet_name == "Transaction History":
                columns_to_format = numeric_format_columns['transaction']
            elif sheet_name == "Hex Dashboard":
                columns_to_format = numeric_format_columns['hex']
            elif sheet_name == "Clients":
                columns_to_format = numeric_format_columns['client']
            elif sheet_name == "OrdersHistory":
                columns_to_format = numeric_format_columns['order']
            else:
                columns_to_format = []

            # Apply formatting for specific columns
            for column_name in columns_to_format:
                if column_name in data.columns:
                    col_index = data.columns.get_loc(column_name)
                    col_letter = chr(65 + col_index)
                    # Format the entire column or specify sufficient rows
                    full_column_range = f"{col_letter}:{col_letter}"
                    
                    try:
                        format_cell_range(worksheet, full_column_range, CellFormat(numberFormat={'type': 'NUMBER', 'pattern': '#,##0'}))
                       # logger.info(f"Formatted column {column_name} in range {full_column_range}")
                    except Exception as format_exception:
                        logger.error(f"Failed to format column {column_name}. Error: {str(format_exception)}")
                else:
                    logger.warning(f"{column_name} not found in DataFrame columns.")
            
            print(f"Data from DataFrame uploaded successfully to sheet '{sheet_name}'!")
            self.is_uploading = False 
                    
        except Exception as e:
            logger.error(f"Error uploading DataFrame to Google Sheets: {str(e)}")
            self.is_uploading = False 

    def load_data_from_sheet(self, sheet_name):
        try:
            # Open the Google Sheet
            spreadsheet = self.client.open_by_key(self.sheet_id)
            try:
                worksheet = spreadsheet.worksheet(sheet_name)
            except gspread.exceptions.WorksheetNotFound:
                logger.error(f"Worksheet '{sheet_name}' not found.")
                return None  # Return None if the sheet does not exist

            # Get all the values in the sheet
            data = worksheet.get_all_values()

            # Convert to DataFrame
            df = pd.DataFrame(data[1:], columns=data[0])  # First row as header

            print(f"Data loaded successfully from sheet '{sheet_name}'!")
            return df  # Return the DataFrame
        except Exception as e:
            logger.error(f"Error loading data from Google Sheets: {str(e)}")
            return None
        
    def delete_all_spreadsheets(self):
        """Delete all spreadsheets in the account."""
        try:
            # List all spreadsheets available in the account
            service = build('drive', 'v3', credentials=self.creds)
            results = service.files().list(q="mimeType='application/vnd.google-apps.spreadsheet'").execute()
            spreadsheets = results.get('files', [])

            if not spreadsheets:
                logger.info("No spreadsheets found.")
                return

            for spreadsheet in spreadsheets:
                spreadsheet_id = spreadsheet['id']
                spreadsheet_name = spreadsheet['name']
                # Delete the spreadsheet
                service.files().delete(fileId=spreadsheet_id).execute()
                logger.info(f"Deleted spreadsheet: {spreadsheet_name} (ID: {spreadsheet_id})")

        except Exception as e:
            logger.error(f"Error deleting spreadsheets: {str(e)}")

  
        

    async def handle_download_google_sheet(self, query, context):
        """Handle the downloading of worksheets filtered for today's data as a single Excel file with specific column names."""
        spreadsheet_name = "Report_" + datetime.now().strftime("%B_%Y")  # Get the correct spreadsheet name
        spreadsheet_id = self.get_spreadsheet_id(spreadsheet_name)

        # Define the new column names for each worksheet
        new_transaction_columns = ['DATE','Account ID','Order Ticket','Client Name','Type','Currency', 'Amount']
        new_hex_columns = ['DATE', 'Bought USDT', 'Sold USDT', 'Net Position']
        new_client_columns = ['Client ID', 'Account ID', 'First Name', 'USDT', 'TOMAN']
        new_order_columns = ['DATE','Account ID', 'Ticket','Client Name', 'Type', 'Currency',
                            'Size', 'ExchangeRate', 'Payable Toman', 'Status', 'Paid', 'Dept']

        if spreadsheet_id:
                while self.is_uploading:
                     await context.bot.send_message(chat_id=query.message.chat.id, text="Wait 15 seconds - Uploading On the process...")
                     time.sleep(15)  # Wait if uploading is in progress
            
                try:
                    # Create a new Excel workbook
                    wb = Workbook()

                    # Get all worksheets from the spreadsheet
                    spreadsheet = self.client.open_by_key(spreadsheet_id)
                    worksheets = spreadsheet.worksheets()
                
                    # Get today's date
                    today_date = datetime.now().date()

                    for worksheet in worksheets:
                        # Download each worksheet data
                        data = self.client.open_by_key(spreadsheet_id).worksheet(worksheet.title).get_all_values()
                        
                        # Prepare a variable to hold filtered data
                        filtered_data = []

                        if worksheet.title == "Hex Dashboard":
                            # Filter based on the 'yyyy-mm-dd' format
                            for row in data:
                                if pd.to_datetime(row[0], errors='coerce').date() == today_date:
                                    filtered_data.append(row)

                            # Add new column names for Hex Dashboard
                            filtered_data.insert(0, new_hex_columns)  # Insert header row

                        elif worksheet.title == "OrdersHistory":
                            # Filter based on the 'yyyy-mm-dd hh:mm:ss' format
                            for row in data:
                                if pd.to_datetime(row[0], errors='coerce').date() == today_date:
                                    filtered_data.append(row)
                            # Insert the column headers for OrdersHistory
                            filtered_data.insert(0, new_order_columns)  # Insert header row

                        elif worksheet.title == "Transaction History":
                            # Filter based on the 'yyyy-mm-dd hh:mm:ss' format
                            for row in data:
                                if pd.to_datetime(row[0], errors='coerce').date() == today_date:
                                    filtered_data.append(row)
                            # Insert the column headers for Transaction History
                            filtered_data.insert(0, new_transaction_columns)  # Use new_order_columns since both have the same structure

                        elif worksheet.title == "Clients":
                            # No filtering needed, include all data
                            filtered_data = data
                            # Add new column names for Clients
                            # filtered_data.insert(0, new_client_columns)  # Insert header row

                        else:
                            # If an unexpected worksheet title is encountered, skip
                            logger.warning(f"Unexpected worksheet title: {worksheet.title}. No data added.")
                            continue

                        # Create a new sheet in the workbook with the same title as the worksheet
                        ws = wb.create_sheet(title=worksheet.title)

                        # Write the filtered data (or full data) to the sheet
                        for row in filtered_data:
                            ws.append(row)  # Append the rows to the workbook

                    # Remove the default sheet created by Workbook if it exists
                    if 'Sheet' in wb.sheetnames:
                        std = wb['Sheet']
                        wb.remove(std)

                    # Save the workbook to a temporary file
                    excel_file_path = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx').name
                    wb.save(excel_file_path)

                    # Create a filename with today's date included
                    today_date_str = datetime.now().strftime("%d") 
                    excel_file_name = f"{spreadsheet_name}_{today_date_str}.xlsx"

                    # Send the file to the user
                    with open(excel_file_path, 'rb') as file:
                        await context.bot.send_document(chat_id=query.message.chat.id, document=file, filename=excel_file_name)

                    os.remove(excel_file_path)  # Clean up the temporary file
                    # await context.bot.send_message(chat_id=query.message.chat.id, text="Today's data from all sheets has been successfully combined into a single Excel file with the correct column names.")

                except Exception as e:
                    logger.error(f"Error during downloading today's data from all worksheets: {str(e)}")
                    await context.bot.send_message(chat_id=query.message.chat.id, text="An error occurred while downloading today's data.")
                  
        else:
            await context.bot.send_message(chat_id=query.message.chat.id, text="Spreadsheet not found.")

    async def ssssshandle_download_google_sheet(self, query, context):
            """Handle the downloading of all worksheets in the Google Sheet as a single Excel file."""
            spreadsheet_name = "Report_" + datetime.now().strftime("%B_%Y")  # Get the correct spreadsheet name
            spreadsheet_id = self.get_spreadsheet_id(spreadsheet_name)

            if spreadsheet_id:
                try:
                    # Create a new Excel workbook
                    wb = Workbook()

                    # Get all worksheets from the spreadsheet
                    spreadsheet = self.client.open_by_key(spreadsheet_id)
                    worksheets = spreadsheet.worksheets()
                    
                    for worksheet in worksheets:
                        # Download each worksheet data
                        data = self.client.open_by_key(spreadsheet_id).worksheet(worksheet.title).get_all_values()
                        
                        # Create a new sheet in the workbook with the same title as the worksheet
                        ws = wb.create_sheet(title=worksheet.title)
                        
                        # Write the data to the sheet
                        for row in data:
                            ws.append(row)  # Append the rows to the workbook

                    # Remove the default sheet created by Workbook
                    if 'Sheet' in wb.sheetnames:
                        std = wb['Sheet']
                        wb.remove(std)

                    # Save the workbook to a temporary file
                    excel_file_path = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx').name
                    wb.save(excel_file_path)

                    # Send the file to the user
                    with open(excel_file_path, 'rb') as file:
                        await context.bot.send_document(chat_id=query.message.chat.id, document=file, filename=os.path.basename(excel_file_path))

                    os.remove(excel_file_path)  # Clean up the temporary file
                    await context.bot.send_message(chat_id=query.message.chat.id, text="All sheets have been successfully combined into a single Excel file.")

                except Exception as e:
                    logger.error(f"Error during downloading all worksheets: {str(e)}")
                    await context.bot.send_message(chat_id=query.message.chat.id, text="An error occurred while downloading the sheets.")
                    
            else:
                await context.bot.send_message(chat_id=query.message.chat.id, text="Spreadsheet not found.")


    






